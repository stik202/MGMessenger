import argparse
import asyncio
import sys
from pathlib import Path
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import select

# Make imports work regardless of current working directory in container.
for candidate in (Path.cwd(), Path("/app"), Path(__file__).resolve().parents[1]):
    if (candidate / "app").exists():
        sys.path.insert(0, str(candidate))
        break

from app.core.security import hash_password
from app.db.models import ChatGroup, GroupMember, Message, User
from app.db.session import AsyncSessionLocal


def sheet_values(path: Path, sheet_name: str | None):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    for row in ws.iter_rows(values_only=True):
        yield ["" if x is None else str(x).strip() for x in row]


async def upsert_user(session, login: str, password: str | None = None):
    login = login.strip()
    if not login:
        return None
    user = await session.scalar(select(User).where(User.login == login))
    if not user:
        user = User(login=login, password_hash=hash_password(password or "1234"))
        session.add(user)
        await session.flush()
    elif password:
        user.password_hash = hash_password(password)
    return user


async def import_auth(session, auth_file: Path, login_col: int, pass_col: int):
    for i, row in enumerate(sheet_values(auth_file, None)):
        if i == 0:
            continue
        login = row[login_col - 1] if len(row) >= login_col else ""
        password = row[pass_col - 1] if len(row) >= pass_col else ""
        if login:
            await upsert_user(session, login, password)


async def import_profiles(session, base_file: Path):
    rows = list(sheet_values(base_file, "Profiles"))
    for i, row in enumerate(rows):
        if i == 0:
            continue
        if not row:
            continue
        login = row[0] if len(row) > 0 else ""
        if not login:
            continue

        user = await upsert_user(session, login)
        user.avatar_url = row[1] if len(row) > 1 else ""
        user.phone = row[2] if len(row) > 2 else ""
        user.email = row[3] if len(row) > 3 else ""
        user.position = row[4] if len(row) > 4 else ""
        user.role = row[5] if len(row) > 5 else "User"
        user.last_name = row[6] if len(row) > 6 else ""
        user.first_name = row[7] if len(row) > 7 else ""
        user.middle_name = row[8] if len(row) > 8 else ""


async def import_groups(session, base_file: Path):
    wb = load_workbook(base_file, data_only=True)
    if "Groups" not in wb.sheetnames:
        return

    ws = wb["Groups"]
    for i, row_data in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue

        row = ["" if x is None else str(x).strip() for x in row_data]
        if not row or not row[0]:
            continue

        gid = row[0]
        name = row[1] if len(row) > 1 else "Группа"
        avatar = row[2] if len(row) > 2 else ""
        members_raw = row[3] if len(row) > 3 else ""
        owner_login = row[4] if len(row) > 4 else ""

        owner = await upsert_user(session, owner_login or "admin")

        existing = await session.scalar(select(ChatGroup).where(ChatGroup.id == UUID(gid)))
        if existing:
            group = existing
            group.name = name
            group.avatar_url = avatar
            group.owner_id = owner.id
            await session.flush()
            await session.execute(GroupMember.__table__.delete().where(GroupMember.group_id == group.id))
        else:
            group = ChatGroup(id=UUID(gid), name=name, avatar_url=avatar, owner_id=owner.id)
            session.add(group)
            await session.flush()

        member_logins = [x.strip() for x in members_raw.split(",") if x.strip()]
        if owner.login not in member_logins:
            member_logins.append(owner.login)

        for login in member_logins:
            member = await upsert_user(session, login)
            session.add(GroupMember(group_id=group.id, user_id=member.id))


async def import_messages(session, base_file: Path):
    wb = load_workbook(base_file, data_only=True)
    if "Messages" not in wb.sheetnames:
        return

    ws = wb["Messages"]
    for i, row_data in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        row = ["" if x is None else str(x).strip() for x in row_data]
        if not row:
            continue

        msg_id = row[0] if len(row) > 0 else ""
        sender_login = row[2] if len(row) > 2 else ""
        receiver = row[4] if len(row) > 4 else ""
        msg_type = row[5] if len(row) > 5 else "private"
        text = row[6] if len(row) > 6 else ""
        file_mime = row[7] if len(row) > 7 else ""
        file_url = row[8] if len(row) > 8 else ""

        if not sender_login:
            continue
        sender = await upsert_user(session, sender_login)

        existing = None
        if msg_id:
            try:
                existing = await session.scalar(select(Message).where(Message.id == UUID(msg_id)))
            except ValueError:
                existing = None
        if existing:
            continue

        message = Message(sender_id=sender.id, text=text, file_mime=file_mime, file_url=file_url)
        if msg_type == "group":
            try:
                message.group_id = UUID(receiver)
            except ValueError:
                continue
        else:
            partner = await upsert_user(session, receiver)
            message.receiver_user_id = partner.id

        session.add(message)


async def main():
    parser = argparse.ArgumentParser(description="Import MGMessenger data from XLSX")
    parser.add_argument("--auth-file", default="/import/Доступы.xlsx")
    parser.add_argument("--base-file", default="/import/MGM base.xlsx")
    parser.add_argument("--auth-login-col", type=int, default=2)
    parser.add_argument("--auth-pass-col", type=int, default=4)
    args = parser.parse_args()

    auth_file = Path(args.auth_file).resolve()
    base_file = Path(args.base_file).resolve()

    async with AsyncSessionLocal() as session:
        await import_auth(session, auth_file, args.auth_login_col, args.auth_pass_col)
        await import_profiles(session, base_file)
        await import_groups(session, base_file)
        await import_messages(session, base_file)
        await session.commit()


if __name__ == "__main__":
    asyncio.run(main())



